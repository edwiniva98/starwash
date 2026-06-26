import os
import logging
import json
import re
from datetime import datetime, timedelta
import httpx
from aiohttp import web
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# ── CONFIG ────────────────────────────────────────────────────────────────────
BOT_TOKEN          = os.environ.get("BOT_TOKEN", "")
OWNER_ID           = int(os.environ.get("OWNER_ID", "0"))
ALLOWED_USERS      = {OWNER_ID, 8838219142}  # Edwin + papá
CLAUDE_KEY         = os.environ.get("ANTHROPIC_API_KEY", "")
FIREBASE_PROJECT   = os.environ.get("FIREBASE_PROJECT_ID", "starwash-cortes")
FIRESTORE_URL      = f"https://firestore.googleapis.com/v1/projects/{FIREBASE_PROJECT}/databases/(default)/documents"
ODOO_URL           = os.environ.get("ODOO_URL", "https://star-wash.odoo.com")
ODOO_DB            = os.environ.get("ODOO_DB", "star-wash")
ODOO_USER          = os.environ.get("ODOO_USER", "starwashtexcoco@hotmail.com")
ODOO_KEY           = os.environ.get("ODOO_KEY", "8f5c64a7d44c7ef76f99545760043dfdda89fa2a")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── FIREBASE ──────────────────────────────────────────────────────────────────
async def get_cortes(limit=5, fecha=None):
    """Obtiene cortes de Firestore."""
    url = f"{FIRESTORE_URL}/cortes"
    params = {"pageSize": limit, "orderBy": "timestamp desc"}
    async with httpx.AsyncClient() as client:
        r = await client.get(url, params=params)
        data = r.json()

    cortes = []
    for doc in data.get("documents", []):
        fields = doc.get("fields", {})
        corte = parse_firestore(fields)
        corte["_id"] = doc["name"].split("/")[-1]
        if fecha:
            if fecha in corte.get("fecha", ""):
                cortes.append(corte)
        else:
            cortes.append(corte)
    return cortes

def parse_firestore(fields):
    """Convierte formato Firestore a dict normal."""
    result = {}
    for key, val in fields.items():
        if "stringValue" in val:
            result[key] = val["stringValue"]
        elif "integerValue" in val:
            result[key] = int(val["integerValue"])
        elif "doubleValue" in val:
            result[key] = float(val["doubleValue"])
        elif "booleanValue" in val:
            result[key] = val["booleanValue"]
        elif "mapValue" in val:
            result[key] = parse_firestore(val["mapValue"].get("fields", {}))
        elif "arrayValue" in val:
            items = val["arrayValue"].get("values", [])
            result[key] = [
                parse_firestore(i.get("mapValue", {}).get("fields", {}))
                if "mapValue" in i else list(i.values())[0]
                for i in items
            ]
        elif "timestampValue" in val:
            result[key] = val["timestampValue"]
    return result

# ── FORMAT ────────────────────────────────────────────────────────────────────
def fmt(n):
    try:
        return f"${float(n):,.2f}"
    except:
        return "$0.00"

def generar_resumen_firebase(corte):
    """Genera resumen completo desde datos de Firebase."""
    v = corte.get("ventas", {})
    a = corte.get("adicionales", {})
    p = corte.get("pagos", {})
    c = corte.get("caja", {})
    t = corte.get("totales", {})
    m = corte.get("maquina", {})
    o = corte.get("otros", {})

    autos_pagados = sum([v.get("autos", 0), v.get("camionetas", 0), v.get("pickups", 0), v.get("express", 0), v.get("fiscalia", 0)])
    total_entraron = autos_pagados + o.get("cortes_taller", 0) + o.get("cortes_ayto", 0) + o.get("cortes_familiar", 0) + o.get("seguro", 0)
    dif_din = t.get("diferencia_din", 0)
    dif_maq = m.get("diferencia", 0)

    estado_din = "✅ Cuadra exacto" if dif_din == 0 else (f"⚠️ Sobra {fmt(abs(dif_din))}" if dif_din > 0 else f"❌ Falta {fmt(abs(dif_din))}")
    estado_maq = "✅ Cuadra exacto" if dif_maq == 0 else (f"⚠️ +{dif_maq} en máquina (posible sin cobrar)" if dif_maq > 0 else f"⚠️ {abs(dif_maq)} menos en máquina")

    msg = f"⚡ *STAR WASH — CORTE DEL DÍA*\n"
    msg += f"📅 {corte.get('fecha', '—')} | 🔖 {corte.get('sesion', '—')}\n"
    msg += f"👤 {corte.get('responsable', '—')}\n\n"

    msg += f"🚗 *VEHÍCULOS*\n"
    msg += f"• Autos: {v.get('autos',0)} | Camionetas: {v.get('camionetas',0)} | Pick-Ups: {v.get('pickups',0)}\n"
    msg += f"• Express: {v.get('express',0)} | Fiscalía: {v.get('fiscalia',0)} | Motos: {v.get('motos',0)}\n"
    msg += f"• *Autos pagados: {autos_pagados}* | Cortesías: {o.get('cortes_taller',0)+o.get('cortes_ayto',0)+o.get('cortes_familiar',0)} (T:{o.get('cortes_taller',0)} A:{o.get('cortes_ayto',0)} F:{o.get('cortes_familiar',0)}) | Seguro: {o.get('seguro',0)}\n"
    msg += f"• *Total entraron: {total_entraron}* | Máquina: {m.get('total_vendidos',0)}\n"
    msg += f"• {estado_maq}\n\n"

    msg += f"💵 *BILLETES EN CAJA: {fmt(c.get('billetes',0))}*\n\n"

    msg += f"💰 *VENTAS ODOO*\n"
    msg += f"• Total tickets: *{fmt(t.get('total_tickets',0))}*\n"
    msg += f"• Efectivo: {fmt(p.get('efectivo',0))} | Tarjeta: {fmt(p.get('tarjeta',0))} | Trans: {fmt(p.get('transferencia',0))}\n"
    msg += f"• Total cobrado: *{fmt(t.get('total_cobrado',0))}*\n\n"

    msg += f"🏦 *CONTROL DE CAJA*\n"
    msg += f"• Apertura: {fmt(c.get('apertura',0))}\n"
    msg += f"• Morralla: {fmt(c.get('morralla',0))} | Billetes: {fmt(c.get('billetes',0))} | Terminal: {fmt(c.get('terminal',0))}\n"
    msg += f"• *Total en caja: {fmt(t.get('total_caja',0))}*\n\n"

    msg += f"📊 *RESULTADO: {estado_din}*\n\n"

    gastos = corte.get("gastos", [])
    if gastos:
        total_g = t.get("total_gastos", 0)
        msg += f"💸 *GASTOS: {fmt(total_g)}*\n"
        for g in gastos:
            if isinstance(g, dict) and g.get("concepto") and g.get("monto", 0) > 0:
                msg += f"• {g['concepto']}: {fmt(g['monto'])}\n"

    svc = corte.get("servicios_manuales", [])
    if svc:
        msg += f"\n✨ *SERVICIOS MANUALES*\n"
        for s in svc:
            if isinstance(s, dict) and s.get("nombre") and s.get("total", 0) > 0:
                msg += f"• {s['nombre']}: {fmt(s['total'])}\n"

    notas = [n for n in corte.get("notas", []) if n and str(n).strip()]
    if notas:
        msg += f"\n📝 *NOTAS*\n"
        for i, n in enumerate(notas, 1):
            msg += f"{i}. {n}\n"

    return msg

# ── DETECCIÓN DE FECHA EN TEXTO ───────────────────────────────────────────────
def detectar_fecha(texto: str) -> str | None:
    """
    Detecta referencias de fecha en el texto del usuario.
    Devuelve fecha en formato YYYY-MM-DD o None.
    """
    hoy = datetime.now()
    texto_lower = texto.lower()

    # Palabras clave relativas
    if any(p in texto_lower for p in ["hoy", "este día", "este dia"]):
        return hoy.strftime("%Y-%m-%d")
    if any(p in texto_lower for p in ["ayer", "el día de ayer", "el dia de ayer"]):
        return (hoy - timedelta(days=1)).strftime("%Y-%m-%d")
    if any(p in texto_lower for p in ["antier", "anteayer"]):
        return (hoy - timedelta(days=2)).strftime("%Y-%m-%d")

    # Días de la semana en español
    dias = {
        "lunes": 0, "martes": 1, "miércoles": 2, "miercoles": 2,
        "jueves": 3, "viernes": 4, "sábado": 5, "sabado": 5, "domingo": 6
    }
    for nombre, num in dias.items():
        if nombre in texto_lower:
            dias_atras = (hoy.weekday() - num) % 7
            if dias_atras == 0:
                dias_atras = 7  # la semana pasada
            return (hoy - timedelta(days=dias_atras)).strftime("%Y-%m-%d")

    # Fecha explícita YYYY-MM-DD
    m = re.search(r"\b(\d{4}-\d{2}-\d{2})\b", texto)
    if m:
        return m.group(1)

    # Fecha DD/MM/YYYY o DD-MM-YYYY
    m = re.search(r"\b(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})\b", texto)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"

    return None

# ── CLAUDE ────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Eres el asistente inteligente del *Autolavado Star Wash*, un negocio de lavado de autos en México.

Tu función es responder preguntas del dueño (Edwin) sobre los cortes del día: ventas, caja, diferencias, gastos, máquina, etc.

Contexto del negocio:
- Categorías de vehículos: Autos ($110), Camionetas ($120), Pick-Ups/SUV ($140), Express ($90), Fiscalía ($80), Motos ($60)
- Adicionales: Tapetes ($40), Motor ($60), Lavado a Mano ($20), Plus+Cera ($20), Pro Cera+Tapetes ($50)
- La "máquina" es el contador automático de vehículos que pasan por el túnel de lavado
- "Corte de caja" = comparar lo que dice Odoo (POS) vs lo que hay en el cajón
- "Diferencia de máquina" = autos que contó la máquina vs autos esperados (pagados + cortesías + seguro - mano - regresos)
- Cortesías: Taller = autos del taller Car Center, Ayuntamiento = vehículos municipales, Familiar = familiares del dueño
- Seguro de lluvia = autos que entran gratis por garantía de lluvia
- "Morralla" = monedas/cambio en caja
- El corte es diario, una sesión por día normalmente

Reglas de respuesta:
- Responde siempre en español, de forma clara y directa
- Usa emojis con moderación para hacer el mensaje legible
- Si te preguntan comparaciones entre fechas, calcula tú mismo los totales
- Si no hay datos para responder, dilo claramente
- Para montos usa formato $X,XXX.XX
- Sé conciso: máximo 5-6 líneas salvo que el detalle sea necesario
- SIEMPRE incluye los billetes en caja (campo caja.billetes) en cualquier resumen o corte completo que presentes. Ponlo así: 💵 Billetes: $X,XXX.XX
- Cuando te pidan "lectura de máquina", "lecturas", "inicial y final" o similar, SIEMPRE muestra la tabla por paquete con este formato exacto (sin tabla markdown, solo texto):
  Paq 1: 12450 → 12478 | 28 autos | $2,240
  Paq 2: 8930 → 8952 | 22 autos | $2,200
  (usa los datos del campo lecturas_maquina del contexto)"""

async def consultar_claude(pregunta: str, cortes: list) -> str:
    """Usa Claude para responder preguntas sobre los cortes con contexto completo."""

    # Construir contexto rico con todo el detalle disponible
    contexto_cortes = []
    for c in cortes:
        t = c.get("totales", {})
        v = c.get("ventas", {})
        a = c.get("adicionales", {})
        p = c.get("pagos", {})
        m = c.get("maquina", {})
        o = c.get("otros", {})
        caja = c.get("caja", {})

        autos_pagados = sum([v.get(k, 0) for k in ["autos", "camionetas", "pickups", "express", "fiscalia"]])

        entrada = {
            "fecha": c.get("fecha"),
            "sesion": c.get("sesion"),
            "responsable": c.get("responsable"),
            "vehiculos": {
                "autos": v.get("autos", 0),
                "camionetas": v.get("camionetas", 0),
                "pickups": v.get("pickups", 0),
                "express": v.get("express", 0),
                "fiscalia": v.get("fiscalia", 0),
                "motos": v.get("motos", 0),
                "total_pagados": autos_pagados,
                "cortesias_taller": o.get("cortes_taller", 0),
                "cortesias_ayto": o.get("cortes_ayto", 0),
                "cortesias_familiar": o.get("cortes_familiar", 0),
                "seguro_lluvia": o.get("seguro", 0),
                "lavados_mano": o.get("lavados_mano", 0),
                "regresos": o.get("regresos", 0),
            },
            "totales": {
                "total_tickets": t.get("total_tickets", 0),
                "total_lavados": t.get("total_lavados", 0),
                "total_adicionales": t.get("total_adicionales", 0),
                "total_cobrado": t.get("total_cobrado", 0),
                "total_caja": t.get("total_caja", 0),
                "total_gastos": t.get("total_gastos", 0),
                "diferencia_caja": t.get("diferencia_din", 0),
            },
            "pagos": {
                "efectivo": p.get("efectivo", 0),
                "tarjeta": p.get("tarjeta", 0),
                "transferencia": p.get("transferencia", 0),
            },
            "caja": {
                "apertura": caja.get("apertura", 0),
                "morralla": caja.get("morralla", 0),
                "billetes": caja.get("billetes", 0),
                "terminal": caja.get("terminal", 0),
            },
            "maquina": {
                "registrada": m.get("total_vendidos", 0),
                "esperada": m.get("esperada", 0),
                "diferencia": m.get("diferencia", 0),
            },
            "gastos": [
                {"concepto": g["concepto"], "monto": g["monto"]}
                for g in c.get("gastos", [])
                if isinstance(g, dict) and g.get("concepto") and g.get("monto", 0) > 0
            ],
            "servicios_manuales": [
                {"nombre": s["nombre"], "precio": s.get("precio",0), "cantidad": s.get("cantidad",1), "total": s.get("total",0)}
                for s in c.get("servicios_manuales", [])
                if isinstance(s, dict) and s.get("nombre")
            ],
            "adicionales": {
                "tapetes_solo": a.get("tapetes_solo", 0),
                "motor": a.get("motor", 0),
                "lavado_mano": a.get("mano", 0),
                "plus_cera": a.get("plus_cera", 0),
                "pro_cera_tapetes": a.get("pro_cera_tapetes", 0),
            },
        }

        # Incluir ventas_detalle y adicionales_detalle si existen (nuevos campos)
        if c.get("ventas_detalle"):
            entrada["ventas_detalle"] = c["ventas_detalle"]
        if c.get("adicionales_detalle"):
            entrada["adicionales_detalle"] = c["adicionales_detalle"]

        # Lecturas iniciales y finales de máquina por paquete
        paquetes = c.get("paquetes_finales", [])
        if paquetes:
            entrada["lecturas_maquina"] = [
                {
                    "paquete": p.get("num", i + 1),
                    "inicio": p.get("inicio", 0),
                    "final": p.get("final", 0),
                    "vendidos": p.get("vendidos", 0),
                    "monto": p.get("monto", 0),
                }
                for i, p in enumerate(paquetes)
                if isinstance(p, dict)
            ]

        notas = [n for n in c.get("notas", []) if n and str(n).strip()]
        if notas:
            entrada["notas"] = notas

        contexto_cortes.append(entrada)

    mensaje_usuario = f"""Datos de los últimos cortes de Star Wash:
{json.dumps(contexto_cortes, ensure_ascii=False, indent=2)}

Pregunta de Edwin: {pregunta}"""

    async with httpx.AsyncClient() as client:
        r = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": CLAUDE_KEY,
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json"
            },
            json={
                "model": "claude-haiku-4-5-20251001",
                "max_tokens": 800,
                "system": SYSTEM_PROMPT,
                "messages": [{"role": "user", "content": mensaje_usuario}]
            },
            timeout=30
        )
        data = r.json()

    if "content" not in data:
        logger.error(f"Claude error: {data}")
        return "❌ No pude consultar los datos en este momento."

    return data["content"][0]["text"]

# ── HANDLERS ──────────────────────────────────────────────────────────────────
async def check_allowed(update: Update) -> bool:
    uid = update.effective_user.id
    if OWNER_ID != 0 and uid not in ALLOWED_USERS:
        await update.message.reply_text("⛔ No tienes acceso a este bot.")
        return False
    return True

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    await update.message.reply_text(
        "👋 Hola! Soy el bot de *Star\\-Wash Cortes*\\.\n\n"
        "Puedo hacer lo siguiente:\n"
        "📊 /ultimo — Ver el último corte guardado\n"
        "📅 /fecha 2026\\-06\\-15 — Ver corte de una fecha\n"
        "📋 /historial — Ver últimos 5 cortes\n"
        "💬 O pregúntame algo natural:\n"
        "  _¿cómo estuvo ayer?_\n"
        "  _¿cuántos autos entraron el viernes?_\n"
        "  _¿cuánto gasté esta semana?_\n\n"
        "📸 También mándame foto de la lectura de máquina\\.",
        parse_mode="MarkdownV2"
    )

async def cmd_ultimo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    msg = await update.message.reply_text("⏳ Buscando el último corte...")
    cortes = await get_cortes(limit=1)
    if not cortes:
        await msg.edit_text("No hay cortes guardados aún.")
        return
    resumen = generar_resumen_firebase(cortes[0])
    await msg.edit_text(resumen, parse_mode="Markdown")

async def cmd_fecha(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    if not context.args:
        await update.message.reply_text("Uso: /fecha 2026-06-15")
        return
    fecha = context.args[0]
    msg = await update.message.reply_text(f"⏳ Buscando corte del {fecha}...")
    cortes = await get_cortes(limit=30, fecha=fecha)
    if not cortes:
        await msg.edit_text(f"No encontré corte para el {fecha}.")
        return
    resumen = generar_resumen_firebase(cortes[0])
    await msg.edit_text(resumen, parse_mode="Markdown")

async def cmd_exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    msg = await update.message.reply_text("⏳ Generando Excel con historial completo...")
    try:
        historial = await get_historial_firestore(limite=500)
        if not historial:
            await msg.edit_text("No hay datos en el historial.")
            return
        
        # Crear Excel en memoria
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Star Wash"
        
        # Encabezados
        headers = ["Fecha", "Día", "Sesión", "Total ($)", "Tickets"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="714B67")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos ordenados por fecha
        historial_sorted = sorted(historial, key=lambda x: x.get("fecha", ""))
        total_general = 0
        for row, d in enumerate(historial_sorted, 2):
            ws.cell(row=row, column=1, value=d.get("fecha", ""))
            ws.cell(row=row, column=2, value=d.get("dia_semana", ""))
            ws.cell(row=row, column=3, value=d.get("sesion", ""))
            ws.cell(row=row, column=4, value=d.get("total", 0))
            ws.cell(row=row, column=5, value=d.get("num_tickets", 0))
            total_general += d.get("total", 0)
            # Colorear sábados y domingos
            dia = d.get("dia_semana", "")
            if dia in ["Sábado", "Domingo"]:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="EAD5F5")
        
        # Fila de totales
        total_row = len(historial_sorted) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=total_general).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=sum(d.get("num_tickets",0) for d in historial_sorted)).font = Font(bold=True)
        
        # Ajustar anchos
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 10
        
        # Guardar en memoria y enviar
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        await update.message.reply_document(
            document=buffer,
            filename="Historial_StarWash.xlsx",
            caption=f"📊 Historial completo Star Wash\n{len(historial_sorted)} días | Total: ${total_general:,.2f}"
        )
        await msg.delete()
        
    except Exception as e:
        logger.error(f"Error exportar: {e}")
        await msg.edit_text(f"Error al exportar: {str(e)}")

async def cmd_sincronizar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    msg = await update.message.reply_text("⏳ Sincronizando historial de Odoo... puede tardar 1-2 minutos.")
    try:
        guardados, errores, total_dias = await sincronizar_historial_odoo()
        texto = f"Sincronizacion completa\n{total_dias} dias procesados\n{guardados} guardados en Firestore\n{errores} errores"
        await msg.edit_text(texto)
    except Exception as e:
        logger.error(f"Error sincronizar: {e}")
        await msg.edit_text(f"❌ Error: {str(e)}")

async def cmd_historial(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    msg = await update.message.reply_text("⏳ Cargando historial...")
    cortes = await get_cortes(limit=5)
    if not cortes:
        await msg.edit_text("No hay cortes guardados.")
        return
    texto = "📋 *ÚLTIMOS CORTES*\n\n"
    for c in cortes:
        t = c.get("totales", {})
        dif = t.get("diferencia_din", 0)
        estado = "✅" if dif == 0 else "⚠️"
        autos = sum([c.get("ventas", {}).get(k, 0) for k in ["autos", "camionetas", "pickups", "express", "fiscalia"]])
        texto += f"{estado} *{c.get('fecha','—')}* — {c.get('sesion','—')}\n"
        texto += f"   🚗 {autos} autos | 💰 {fmt(t.get('total_tickets',0))} | 💸 {fmt(t.get('total_gastos',0))} gastos\n\n"
    await msg.edit_text(texto, parse_mode="Markdown")

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    caption = update.message.caption or "📸 Lectura de máquina"
    sender = update.effective_user.first_name
    await update.message.reply_text("✅ Foto recibida.")
    if update.effective_user.id != OWNER_ID:
        photo = update.message.photo[-1]
        await context.bot.send_photo(
            chat_id=OWNER_ID,
            photo=photo.file_id,
            caption=f"📸 Lectura de máquina — {sender}\n{caption}"
        )


# ── HISTORIAL ODOO ────────────────────────────────────────────────────────────
async def get_historial_odoo(dias=None):
    """Obtiene historial completo de ventas POS desde Odoo usando órdenes."""
    async with httpx.AsyncClient(timeout=120) as client:
        uid = await odoo_uid(client)
        
        # Usar órdenes en lugar de sesiones — tienen fecha correcta siempre
        ordenes = await odoo_call(client, uid, 'pos.order', 'search_read',
            [[['state', 'in', ['done', 'invoiced']]]],
            {
                'fields': ['name', 'date_order', 'amount_total', 'session_id'],
                'order': 'date_order desc',
                'limit': 50000
            }
        )
        
        # Agrupar por día
        from collections import defaultdict
        por_dia = defaultdict(lambda: {'total': 0, 'ordenes': 0, 'sesion': ''})
        
        for o in ordenes:
            fecha = (o.get('date_order') or '')[:10]
            if not fecha:
                continue
            por_dia[fecha]['total'] += o.get('amount_total', 0)
            por_dia[fecha]['ordenes'] += 1
            if o.get('session_id'):
                por_dia[fecha]['sesion'] = o['session_id'][1] if isinstance(o['session_id'], list) else str(o['session_id'])
        
        resumen = [
            {
                'fecha': fecha,
                'sesion': datos['sesion'],
                'total': round(datos['total'], 2),
                'num_tickets': datos['ordenes'],
            }
            for fecha, datos in sorted(por_dia.items(), reverse=True)
        ]
        
    return resumen

# ── SINCRONIZACIÓN ODOO → FIRESTORE ──────────────────────────────────────────
async def sincronizar_historial_odoo():
    """Descarga todas las órdenes de Odoo y las guarda en Firestore agrupadas por día."""
    logger.info("Iniciando sincronización Odoo → Firestore...")
    
    async with httpx.AsyncClient(timeout=300) as client:
        uid = await odoo_uid(client)
        
        # Traer todas las órdenes paginando
        ordenes = []
        offset = 0
        batch_size = 5000
        while True:
            batch = await odoo_call(client, uid, 'pos.order', 'search_read',
                [[['state', 'in', ['done', 'invoiced']]]],
                {
                    'fields': ['name', 'date_order', 'amount_total', 'session_id'],
                    'order': 'date_order asc',
                    'limit': batch_size,
                    'offset': offset
                }
            )
            if not batch:
                break
            ordenes.extend(batch)
            offset += batch_size
            logger.info(f"Batch {offset//batch_size}: {len(batch)} ordenes")
            if len(batch) < batch_size:
                break
        
        logger.info(f"Total ordenes: {len(ordenes)}")
        
        # Agrupar por día
        from collections import defaultdict
        por_dia = defaultdict(lambda: {'total': 0, 'tickets': 0, 'sesion': '', 'fecha': ''})
        
        for o in ordenes:
            fecha = (o.get('date_order') or '')[:10]
            if not fecha:
                continue
            por_dia[fecha]['total'] += o.get('amount_total', 0)
            por_dia[fecha]['tickets'] += 1
            por_dia[fecha]['fecha'] = fecha
            if o.get('session_id'):
                por_dia[fecha]['sesion'] = o['session_id'][1] if isinstance(o['session_id'], list) else str(o['session_id'])
        
        # Guardar cada día en Firestore colección 'historial_odoo'
        from datetime import datetime as dt_sync
        DIAS_ES = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
        MESES_ES = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
        guardados = 0
        errores = 0
        for fecha, datos in por_dia.items():
            doc_id = fecha  # YYYY-MM-DD como ID del documento
            firestore_url = f"{FIRESTORE_URL}/historial_odoo/{doc_id}"
            try:
                fecha_dt = dt_sync.strptime(fecha, "%Y-%m-%d")
                dia_semana = DIAS_ES[fecha_dt.weekday()]
                fecha_legible = f"{dia_semana} {fecha_dt.day} de {MESES_ES[fecha_dt.month]} de {fecha_dt.year}"
            except:
                dia_semana = ""
                fecha_legible = fecha
            
            body = {
                "fields": {
                    "fecha":         {"stringValue": fecha},
                    "fecha_legible": {"stringValue": fecha_legible},
                    "dia_semana":    {"stringValue": dia_semana},
                    "sesion":        {"stringValue": datos['sesion']},
                    "total":         {"doubleValue": round(datos['total'], 2)},
                    "num_tickets":   {"integerValue": str(datos['tickets'])},
                    "fuente":        {"stringValue": "odoo"},
                    "sincronizado":  {"stringValue": datetime.now().isoformat()},
                }
            }
            
            try:
                r = await client.patch(firestore_url, json=body)
                if r.status_code in [200, 201]:
                    guardados += 1
                else:
                    errores += 1
                    logger.error(f"Error guardando {fecha}: {r.status_code}")
            except Exception as e:
                errores += 1
                logger.error(f"Error {fecha}: {e}")
        
        logger.info(f"Sincronización completa: {guardados} días guardados, {errores} errores")
        return guardados, errores, len(por_dia)

async def get_historial_firestore(limite=500):
    """Lee el historial de Firestore (más rápido que Odoo)."""
    url = f"{FIRESTORE_URL}/historial_odoo"
    params = {"pageSize": limite}
    async with httpx.AsyncClient() as client:
        r = await client.get(url, params=params)
        data = r.json()
    
    docs = data.get("documents", [])
    from datetime import datetime as dt2
    DIAS_SEM = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
    historial = []
    for doc in docs:
        fields = doc.get("fields", {})
        fecha_str = fields.get("fecha", {}).get("stringValue", "")
        try:
            dia_sem = DIAS_SEM[dt2.strptime(fecha_str, "%Y-%m-%d").weekday()] if fecha_str else ""
        except:
            dia_sem = ""
        historial.append({
            "fecha":         fecha_str,
            "fecha_legible": fields.get("fecha_legible", {}).get("stringValue", "") or dia_sem + " " + fecha_str,
            "dia_semana":    fields.get("dia_semana", {}).get("stringValue", "") or dia_sem,
            "sesion":        fields.get("sesion", {}).get("stringValue", ""),
            "total":         float(fields.get("total", {}).get("doubleValue", 0) or fields.get("total", {}).get("integerValue", 0) or 0),
            "num_tickets":   int(fields.get("num_tickets", {}).get("integerValue", 0) or 0),
        })
    return historial

# ── AUDITORÍA AUTOMÁTICA ──────────────────────────────────────────────────────
async def auditar_corte(corte: dict, historial: list) -> str:
    """Genera reporte de auditoría comparando el corte vs historial."""
    alertas = []
    info = []

    t = corte.get("totales", {})
    v = corte.get("ventas", {})
    c = corte.get("caja", {})
    m = corte.get("maquina", {})
    o = corte.get("otros", {})

    autos_pagados = sum([v.get(k, 0) for k in ["autos", "camionetas", "pickups", "express", "fiscalia"]])
    dif_din = t.get("diferencia_din", 0)
    dif_maq = m.get("diferencia", 0)
    billetes = c.get("billetes", 0)
    efectivo_odoo = corte.get("pagos", {}).get("efectivo", 0)
    gastos = t.get("total_gastos", 0)

    # 1. Diferencia de caja
    if dif_din != 0:
        if abs(dif_din) >= 1000:
            alertas.append(f"🚨 Diferencia de caja: {fmt(abs(dif_din))} ({'sobrante' if dif_din > 0 else 'faltante'}) — revisar conteo de efectivo")
        elif abs(dif_din) > 0:
            alertas.append(f"⚠️ Diferencia de caja menor: {fmt(abs(dif_din))} ({'sobrante' if dif_din > 0 else 'faltante'})")
    else:
        info.append("✅ Caja cuadra exacto")

    # 2. Diferencia de máquina
    if dif_maq != 0:
        causa = ""
        lavm = o.get("lavados_mano", 0)
        reg = o.get("regresos", 0)
        if lavm > 0:
            causa = f" — hay {lavm} lavado(s) a mano que no pasan por máquina"
        elif reg > 0:
            causa = f" — hay {reg} regreso(s) registrado(s)"
        if dif_maq > 0:
            alertas.append(f"⚠️ Máquina marcó {dif_maq} auto(s) de más{causa} — posible auto sin cobrar")
        else:
            alertas.append(f"⚠️ Máquina marcó {abs(dif_maq)} auto(s) de menos{causa}")
    else:
        info.append("✅ Máquina cuadra exacto")

    # 3. Billetes vs efectivo Odoo
    if efectivo_odoo > 0 and billetes > 0:
        morralla = c.get("morralla", 0)
        apertura = c.get("apertura", 0)
        efectivo_real = billetes + morralla - apertura
        dif_efe = efectivo_real - efectivo_odoo
        if abs(dif_efe) > 100:
            alertas.append(f"⚠️ Billetes+Morralla-Apertura ({fmt(efectivo_real)}) vs Efectivo Odoo ({fmt(efectivo_odoo)}) — diferencia {fmt(abs(dif_efe))}")

    # 4. Comparar vs historial (si hay suficientes cortes)
    if len(historial) >= 2:
        hist_autos = [sum([h.get("ventas", {}).get(k, 0) for k in ["autos", "camionetas", "pickups", "express", "fiscalia"]]) for h in historial]
        hist_gastos = [h.get("totales", {}).get("total_gastos", 0) for h in historial]
        hist_tickets = [h.get("totales", {}).get("total_tickets", 0) for h in historial]

        prom_autos = sum(hist_autos) / len(hist_autos)
        prom_gastos = sum(hist_gastos) / len(hist_gastos)
        prom_tickets = sum(hist_tickets) / len(hist_tickets)

        # Autos bajos
        if prom_autos > 0 and autos_pagados < prom_autos * 0.75:
            alertas.append(f"📉 Autos pagados ({autos_pagados}) están {round((1 - autos_pagados/prom_autos)*100)}% por debajo del promedio ({round(prom_autos)})")

        # Gastos inusuales
        if prom_gastos > 0 and gastos > prom_gastos * 1.30:
            alertas.append(f"💸 Gastos ({fmt(gastos)}) superan 30% del promedio histórico ({fmt(prom_gastos)}) — revisar")

        # Ventas bajas
        tickets = t.get("total_tickets", 0)
        if prom_tickets > 0 and tickets < prom_tickets * 0.75:
            alertas.append(f"📉 Ventas ({fmt(tickets)}) están {round((1 - tickets/prom_tickets)*100)}% por debajo del promedio ({fmt(prom_tickets)})")

        info.append(f"📊 Promedios históricos ({len(historial)} días): {round(prom_autos)} autos | {fmt(prom_tickets)} ventas | {fmt(prom_gastos)} gastos")

    # Construir mensaje
    fecha = corte.get("fecha", "—")
    sesion = corte.get("sesion", "—")
    msg = f"🔍 *AUDITORÍA — {fecha} | {sesion}*\n\n"

    if alertas:
        msg += "\n".join(alertas)
    else:
        msg += "\n".join(info)
        msg += "\n\n✅ Todo en orden, sin anomalías detectadas."

    if alertas and info:
        msg += "\n\n" + "\n".join(info)

    return msg

PALABRAS_HISTORICO = [
    'mejor día', 'mejor dia', 'peor día', 'peor dia',
    'más vendido', 'mas vendido', 'mayor venta', 'menor venta',
    'semana pasada', 'mes pasado', 'último mes', 'ultimos dias',
    'promedio', 'tendencia', 'comparar', 'histórico', 'historico',
    'cuánto hemos vendido', 'cuanto hemos vendido',
    'ranking', 'top', 'los mejores', 'los peores',
    'este mes', 'esta semana', 'del mes', 'de la semana',
]

PALABRAS_CORTE_COMPLETO = [
    "dame el corte", "mándame el corte", "mandame el corte",
    "el corte de", "corte del día", "corte del dia",
    "ver corte", "muéstrame el corte", "muestrame el corte",
    "corte completo", "resumen del día", "resumen del dia",
    "cómo estuvo", "como estuvo", "qué tal estuvo", "que tal estuvo",
]

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await check_allowed(update): return
    texto = update.message.text
    texto_lower = texto.lower()
    msg = await update.message.reply_text("⏳ Consultando datos...")

    try:
        fecha_detectada = detectar_fecha(texto)

        # Si piden el corte completo → usar generar_resumen_firebase (formato fijo, siempre incluye billetes)
        es_corte_completo = any(p in texto_lower for p in PALABRAS_CORTE_COMPLETO)

        if es_corte_completo:
            if fecha_detectada:
                cortes = await get_cortes(limit=30, fecha=fecha_detectada)
            else:
                cortes = await get_cortes(limit=1)
            if not cortes:
                await msg.edit_text("No encontré corte para ese día.")
                return
            resumen = generar_resumen_firebase(cortes[0])
            await msg.edit_text(resumen, parse_mode="Markdown")
            return

        # Preguntas históricas → usar historial completo
        # Detectar si menciona mes/año pasado o fecha antigua
        meses = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto',
                 'septiembre','octubre','noviembre','diciembre']
        menciona_mes = any(m in texto_lower for m in meses)
        menciona_anio_pasado = '2025' in texto or '2024' in texto or '2026' in texto
        palabras_ranking = ['top ', 'mejores dias', 'mejores días', 'peores dias', 'peores días',
                           'ranking', 'historial', 'desde siempre', 'todos los dias', 'todos los meses',
                           'cuanto hemos', 'cuánto hemos', 'en total', 'historico', 'histórico']
        es_ranking = any(p in texto_lower for p in palabras_ranking)
        es_historico = any(p in texto_lower for p in PALABRAS_HISTORICO) or menciona_mes or menciona_anio_pasado or es_ranking

        if es_historico:
            await msg.edit_text("⏳ Consultando historial completo de Odoo...")
            try:
                # Intentar desde Firestore primero (más rápido)
                historial_odoo = await get_historial_firestore()
                if len(historial_odoo) < 10:
                    # Si no hay suficiente en Firestore, ir a Odoo
                    historial_odoo = await get_historial_odoo()
                prompt_hist = f"""Eres el asistente del Autolavado Star Wash en Texcoco, México. El dueño (Edwin) pregunta sobre el historial de ventas.

Tienes acceso al historial COMPLETO del negocio con {len(historial_odoo)} días de datos desde julio 2025.
Cada registro tiene: fecha (YYYY-MM-DD), total (ventas del día en pesos MXN), num_tickets (número de órdenes).

Historial completo:
{json.dumps(historial_odoo, ensure_ascii=False, indent=2)}

Pregunta de Edwin: {texto}

INSTRUCCIONES:
- Responde en español con emojis
- Usa los datos exactos del historial — no inventes ni estimes
- Sé específico con fechas completas (día, mes, año)
- Formato de dinero: $XX,XXX.00
- Si piden ranking o top, ordena correctamente por el criterio pedido
- No digas "últimos 90 días" — tienes el historial completo"""
                async with httpx.AsyncClient() as client_ai:
                    r = await client_ai.post(
                        "https://api.anthropic.com/v1/messages",
                        headers={"x-api-key": CLAUDE_KEY, "anthropic-version": "2023-06-01", "Content-Type": "application/json"},
                        json={"model": "claude-haiku-4-5-20251001", "max_tokens": 1000, "messages": [{"role": "user", "content": prompt_hist}]},
                        timeout=60
                    )
                    data = r.json()
                respuesta = data["content"][0]["text"]
                respuesta = respuesta.replace("# ", "").replace("## ", "").replace("### ", "")
                await msg.edit_text(respuesta, parse_mode="Markdown")
                return
            except Exception as e:
                logger.error(f"Error historial Odoo: {e}")
                await msg.edit_text(f"❌ Error consultando historial: {str(e)}")
                return

        # Preguntas analíticas → Claude con contexto completo
        if fecha_detectada:
            cortes_fecha = await get_cortes(limit=30, fecha=fecha_detectada)
            cortes_recientes = await get_cortes(limit=5)
            ids_vistos = {c["_id"] for c in cortes_fecha}
            cortes = cortes_fecha + [c for c in cortes_recientes if c["_id"] not in ids_vistos]
        else:
            cortes = await get_cortes(limit=10)

        if not cortes:
            await msg.edit_text("No hay cortes guardados aún para responder tu pregunta.")
            return

        respuesta = await consultar_claude(texto, cortes)
        # Limpiar formato incompatible con Telegram Markdown
        respuesta = respuesta.replace("# ", "").replace("## ", "").replace("### ", "")
        await msg.edit_text(respuesta, parse_mode="Markdown")

    except Exception as e:
        logger.error(f"Error en handle_text: {e}")
        await msg.edit_text(f"❌ Error consultando datos: {str(e)}")

# ── MAIN ──────────────────────────────────────────────────────────────────────
# ── WEBHOOK HTTP para recibir cortes desde el HTML ────────────────────────────
_bot_app = None  # referencia global al bot

# ── ODOO PROXY ───────────────────────────────────────────────────────────────
async def odoo_uid(client):
    """Obtiene el UID usando XML-RPC con API key."""
    import xmlrpc.client
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_KEY, {})
    logger.info(f"Odoo XML-RPC uid={uid}")
    if not uid:
        raise Exception("Odoo XML-RPC auth failed")
    return uid

async def odoo_call(client, uid, model, method, args=[], kwargs={}):
    """Hace una llamada XML-RPC a Odoo."""
    import xmlrpc.client
    models = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    result = models.execute_kw(ODOO_DB, uid, ODOO_KEY, model, method, args, kwargs)
    return result

async def handle_odoo_fields(request):
    """Endpoint temporal para ver campos de pos.session."""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            uid = await odoo_uid(client)
            fields = await odoo_call(client, uid, 'pos.session', 'fields_get', [], {'attributes': ['string', 'type']})
            # Filtrar campos relevantes de caja/efectivo
            relevant = {k: v for k, v in fields.items() if any(w in k.lower() for w in ['cash', 'total', 'amount', 'register', 'balance'])}
        return web.Response(text=json.dumps(relevant, indent=2), headers={'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'})
    except Exception as e:
        return web.Response(text=json.dumps({"error": str(e)}), status=500, headers={'Access-Control-Allow-Origin': '*'})

async def handle_odoo_sesiones(request):
    """Devuelve las últimas 5 sesiones POS de Odoo."""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            uid = await odoo_uid(client)
            from datetime import datetime, timedelta
            hace_7_dias = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d 00:00:00')
            # Buscar sesiones cerradas de los últimos 7 días + sesión abierta actual
            sesiones_cerradas = await odoo_call(client, uid, 'pos.session', 'search_read',
                [[['state', '=', 'closed'], ['stop_at', '>=', hace_7_dias]]],
                {
                    'fields': ['id', 'name', 'start_at', 'stop_at', 'state', 'total_payments_amount', 'cash_register_balance_start', 'cash_register_difference'],
                    'order': 'stop_at desc',
                    'limit': 5
                }
            )
            sesiones_abiertas = await odoo_call(client, uid, 'pos.session', 'search_read',
                [[['state', '=', 'opened']]],
                {
                    'fields': ['id', 'name', 'start_at', 'stop_at', 'state', 'total_payments_amount', 'cash_register_balance_start', 'cash_register_difference'],
                    'limit': 1
                }
            )
            sesiones = sesiones_abiertas + sesiones_cerradas
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET, OPTIONS',
            'Content-Type': 'application/json'
        }
        return web.Response(text=json.dumps(sesiones), headers=headers)
    except Exception as e:
        logger.error(f"Error odoo_sesiones: {e}")
        return web.Response(text=json.dumps({"error": str(e)}), status=500,
                           headers={'Access-Control-Allow-Origin': '*'})

async def handle_odoo_sesion_detalle(request):
    """Devuelve el detalle completo de una sesión POS."""
    try:
        session_id = int(request.match_info['session_id'])
        async with httpx.AsyncClient(timeout=30) as client:
            uid = await odoo_uid(client)

            # Sesión
            sesiones = await odoo_call(client, uid, 'pos.session', 'search_read',
                [[['id', '=', session_id]]],
                {'fields': ['name', 'start_at', 'stop_at', 'state', 'total_payments_amount', 'cash_register_balance_start', 'cash_register_difference'], 'limit': 1}
            )
            sesion = sesiones[0] if sesiones else {}

            # Pagos agrupados
            pagos = await odoo_call(client, uid, 'pos.payment', 'search_read',
                [[['session_id', '=', session_id]]],
                {'fields': ['amount', 'payment_method_id'], 'limit': 1000}
            )

            # Líneas de venta
            lineas = await odoo_call(client, uid, 'pos.order.line', 'search_read',
                [[['order_id.session_id', '=', session_id]]],
                {'fields': ['product_id', 'qty', 'price_unit', 'price_subtotal_incl'], 'limit': 2000}
            )

        result = {
            'sesion': sesion,
            'pagos': pagos,
            'lineas': lineas,
        }
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET, OPTIONS',
            'Content-Type': 'application/json'
        }
        return web.Response(text=json.dumps(result, default=str), headers=headers)
    except Exception as e:
        logger.error(f"Error odoo_detalle: {e}")
        return web.Response(text=json.dumps({"error": str(e)}), status=500,
                           headers={'Access-Control-Allow-Origin': '*'})

async def handle_options(request):
    """Maneja preflight CORS."""
    return web.Response(headers={
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type',
    })

async def handle_corte_nuevo(request):
    """Recibe un corte guardado desde el HTML y manda auditoría por Telegram."""
    try:
        corte = await request.json()
        # Obtener historial para comparar (excluir el corte actual)
        historial = await get_cortes(limit=10)
        fecha_actual = corte.get("fecha", "")
        historial = [c for c in historial if c.get("fecha") != fecha_actual]

        auditoria_msg = await auditar_corte(corte, historial)

        if _bot_app:
            for uid in ALLOWED_USERS:
                await _bot_app.bot.send_message(
                    chat_id=uid,
                    text=auditoria_msg,
                    parse_mode="Markdown"
                )
        return web.Response(text="ok")
    except Exception as e:
        logger.error(f"Error en handle_corte_nuevo: {e}")
        return web.Response(text=f"error: {e}", status=500)

async def run_web_server():
    """Servidor HTTP para recibir notificaciones del HTML."""
    server = web.Application()
    server.router.add_post("/corte-nuevo", handle_corte_nuevo)
    server.router.add_get("/odoo/sesiones", handle_odoo_sesiones)
    server.router.add_get("/odoo/fields", handle_odoo_fields)
    server.router.add_get("/odoo/sesion/{session_id}", handle_odoo_sesion_detalle)
    server.router.add_route("OPTIONS", "/{path_info:.*}", handle_options)
    runner = web.AppRunner(server)
    await runner.setup()
    port = int(os.environ.get("PORT", 8080))
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logger.info(f"Servidor HTTP en puerto {port}")

def main():
    global _bot_app
    app = Application.builder().token(BOT_TOKEN).build()
    _bot_app = app
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("ultimo", cmd_ultimo))
    app.add_handler(CommandHandler("fecha", cmd_fecha))
    app.add_handler(CommandHandler("historial", cmd_historial))
    app.add_handler(CommandHandler("sincronizar", cmd_sincronizar))
    app.add_handler(CommandHandler("exportar", cmd_exportar))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    import asyncio

    async def run_all():
        await run_web_server()
        await app.initialize()
        await app.start()
        await app.updater.start_polling(drop_pending_updates=True)
        logger.info("Bot iniciado con servidor HTTP...")
        try:
            await asyncio.Event().wait()
        finally:
            await app.updater.stop()
            await app.stop()
            await app.shutdown()

    asyncio.run(run_all())

if __name__ == "__main__":
    main()
